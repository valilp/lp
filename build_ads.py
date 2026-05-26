import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Golden Route - Google Ads"

# ── Colors & styles ──────────────────────────────────────────────────────────
PINK  = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
BLUE  = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
ALT   = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

thin   = Side(border_style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(cell, text, fill=None):
    cell.value = text
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    if fill:
        cell.fill = fill

def data(cell, value, align="left", wrap=False, fill=None):
    cell.value = value
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = BORDER
    if fill:
        cell.fill = fill

# ── Row 1: Headers ───────────────────────────────────────────────────────────
# A: headline  B: wc/30  C: descriptions  D: wc/90  E: Path1  F: wc/15  G: Path2  H: wc/15
hdr(ws["A1"], "見出し\nheadline",       PINK)
hdr(ws["B1"], "文字数\nword count")
hdr(ws["C1"], "説明文\ndescriptions",  PINK)
hdr(ws["D1"], "文字数\nword count")
hdr(ws["E1"], "URL\nPath 1",           BLUE)
hdr(ws["F1"], "文字数\nword count")
hdr(ws["G1"], "URL\nPath 2",           BLUE)
hdr(ws["H1"], "文字数\nword count")

# ── Content ───────────────────────────────────────────────────────────────────
# 15 headlines ≤ 30 chars each
headlines = [
    "Tokyo to Osaka by Shinkansen",       # 28
    "Japan's Golden Route by Train",      # 29
    "Explore Japan's Golden Route",       # 28
    "See Mt. Fuji from the Train",        # 27
    "Tokyo, Kyoto & Osaka by Rail",       # 28
    "Discover Kyoto's Temples",            # 24
    "Shinkansen Pass for Japan",           # 25
    "Golden Route Shinkansen Pass",        # 28
    "Journey Through Historic Japan",     # 30
    "Nara, Kyoto & Tokyo by Rail",        # 27
    "Ride Japan's Bullet Train",          # 25
    "From Tokyo to Hiroshima",             # 23
    "Scenic Views Across Japan",           # 25
    "Travel Japan by Bullet Train",       # 28
    "Book Your Japan Rail Pass",           # 25
]

# 4 descriptions ≤ 90 chars each, on rows 2, 4, 6, 8 (headline indices 0, 2, 4, 6)
descriptions = {
    0: "Plan your trip to Japan. Discover Tokyo, Mt. Fuji, Kyoto, and Osaka by Shinkansen.",  # 82
    2: "From Tokyo to Osaka, the Golden Route takes you through Kyoto, Nara, and Mt. Fuji.",  # 82
    4: "See ancient temples in Kyoto, ride the Shinkansen past Mt. Fuji, and reach Osaka.",   # 81
    6: "Travel Japan's most iconic route. Get your Shinkansen pass and explore more cities.", # 83
}

# URL paths (pre-filled in row 2 only)
PATH1 = "/GoldenRoute"   # 12 chars
PATH2 = "/Itineraries"   # 12 chars

# ── Data rows 2–16 ────────────────────────────────────────────────────────────
for i, headline in enumerate(headlines):
    r   = i + 2                          # spreadsheet row
    bg  = ALT if i % 2 == 0 else None   # alternating row tint

    data(ws[f"A{r}"], headline,                                   fill=bg)
    data(ws[f"B{r}"], f'=LEN(A{r})&"/30"', align="center",       fill=bg)

    if i in descriptions:
        data(ws[f"C{r}"], descriptions[i], wrap=True,             fill=bg)
    else:
        ws[f"C{r}"].border = BORDER
        if bg: ws[f"C{r}"].fill = bg

    data(ws[f"D{r}"], f'=LEN(C{r})&"/90"', align="center",       fill=bg)

    if r == 2:
        data(ws[f"E{r}"], PATH1,                                  fill=bg)
        data(ws[f"G{r}"], PATH2,                                  fill=bg)
    else:
        ws[f"E{r}"].border = BORDER
        ws[f"G{r}"].border = BORDER
        if bg:
            ws[f"E{r}"].fill = bg
            ws[f"G{r}"].fill = bg

    data(ws[f"F{r}"], f'=LEN(E{r})&"/15"', align="center",       fill=bg)
    data(ws[f"H{r}"], f'=LEN(G{r})&"/15"', align="center",       fill=bg)

# ── Layout ────────────────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 13
ws.column_dimensions["C"].width = 54
ws.column_dimensions["D"].width = 13
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 13
ws.column_dimensions["G"].width = 16
ws.column_dimensions["H"].width = 13

ws.row_dimensions[1].height = 44
for r in range(2, 17):
    ws.row_dimensions[r].height = 22

out = "/home/user/lp/JR_Central_GoldenRoute_GoogleAds.xlsx"
wb.save(out)
print(f"Saved → {out}")

# ── Quick sanity-check ────────────────────────────────────────────────────────
print("\nHeadline char counts:")
for h in headlines:
    flag = "✓" if len(h) <= 30 else "✗ OVER"
    print(f"  {len(h):2d}/30 {flag}  {h}")

print("\nDescription char counts:")
for idx, d in descriptions.items():
    flag = "✓" if len(d) <= 90 else "✗ OVER"
    print(f"  {len(d):2d}/90 {flag}  {d}")

print(f"\nPath 1: {len(PATH1)}/15  {PATH1}")
print(f"Path 2: {len(PATH2)}/15  {PATH2}")
